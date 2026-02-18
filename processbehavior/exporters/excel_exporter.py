"""
Excel export functionality for AnalysisResult.

This module provides the ExcelExporter class which handles all Excel export
operations, keeping the AnalysisResult class focused on data access.
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

import pandas as pd

if TYPE_CHECKING:
    from ..analysis_result import AnalysisResult

logger = logging.getLogger(__name__)

# Standard SPC chart type names
STANDARD_CHART_NAMES = {'Xbar', 'S', 'XmR', 'R'}


class ExcelExporter:
    """
    Export AnalysisResult to Excel workbook with multiple tabs.

    This class encapsulates all Excel export functionality, providing a clean
    separation of concerns from the AnalysisResult data container.

    Parameters
    ----------
    result : AnalysisResult
        The analysis result to export

    Examples
    --------
    Direct usage (prefer result.to_excel() for convenience):

    >>> from processbehavior.exporters import ExcelExporter
    >>> exporter = ExcelExporter(result)
    >>> exporter.export('analysis.xlsx')

    Via AnalysisResult delegation:

    >>> result.to_excel('analysis.xlsx')
    """

    def __init__(self, result: AnalysisResult):
        """Initialize exporter with an AnalysisResult."""
        self.result = result

    def export(
        self,
        filepath: str,
        include_summary: bool = True,
        include_charts: bool = True,
        include_residuals: bool = True,
        include_effects: bool = True,
        include_interactions: bool = True,
        include_full_dataset: bool = False,
        format_cells: bool = True,
        include_chart_images: bool = True,
        export_html: bool = True
    ) -> None:
        """
        Export analysis results to Excel with each component on a separate tab.

        Creates a multi-sheet Excel workbook with organized analysis results:
        - Summary: Analysis metadata, SDS info, signal counts
        - Charts: One tab per chart (Xbar, S, stratified XmR, etc.)
        - Residuals: R1-R5 variance decomposition (if available)
        - Effects: Main effects (if calculated)
        - Interactions: Interaction terms (if calculated)
        - Full_Dataset: Complete analysis dataset (optional)

        Parameters
        ----------
        filepath : str
            Output Excel file path (e.g., 'analysis.xlsx')
        include_summary : bool, default=True
            Include summary tab with analysis metadata
        include_charts : bool, default=True
            Include tabs for each chart
        include_residuals : bool, default=True
            Include residuals tab if available
        include_effects : bool, default=True
            Include effects tab if available
        include_interactions : bool, default=True
            Include interactions tab if available
        include_full_dataset : bool, default=False
            Include complete analysis dataset (can be large)
        format_cells : bool, default=True
            Apply formatting (bold headers, auto-width, freeze panes)
        include_chart_images : bool, default=True
            Include visual charts tab with embedded chart images (requires plotly)
        export_html : bool, default=True
            Export companion interactive HTML files alongside Excel file

        Returns
        -------
        None
            File is written to disk

        Examples
        --------
        Export with default settings (all available data except full dataset):

        >>> exporter.export('my_analysis.xlsx')

        Export with full dataset included:

        >>> exporter.export('complete_report.xlsx', include_full_dataset=True)

        Export only charts and summary (minimal):

        >>> exporter.export('charts_only.xlsx',
        ...                 include_residuals=False,
        ...                 include_effects=False,
        ...                 include_interactions=False)

        Notes
        -----
        - Tab names are limited to 31 characters (Excel limitation)
        - Chart tabs are prefixed with 'Chart_' for clarity
        - Stratified charts use format 'XmR_{group_name}'
        - Summary tab includes SDS info and signal counts
        - Formatting includes frozen headers and auto-sized columns
        - Visual_Charts tab includes embedded images (requires Chrome for kaleido)
        - Interactive HTML files are exported alongside Excel for full interactivity
        - HTML charts support zoom, pan, and hover tooltips
        """
        # Create Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:

            # 1. Summary Tab
            if include_summary:
                self._write_summary_tab(writer, format_cells)

            # 2. Chart Tabs
            if include_charts:
                self._write_chart_tabs(writer, format_cells)

            # 3. Residuals Tab
            if include_residuals and self.result.has_residuals:
                self._write_residuals_tab(writer, format_cells)

            # 4. Effects Tab
            if include_effects and self.result.has_effects:
                self._write_effects_tab(writer, format_cells)

            # 5. Interactions Tab
            if include_interactions and self.result.has_interactions:
                self._write_interactions_tab(writer, format_cells)

            # 6. Full Dataset Tab (optional - can be large)
            if include_full_dataset:
                self._write_full_dataset_tab(writer, format_cells)

            # 7. Visual Charts Tab (with embedded images)
            if include_chart_images:
                self._write_visual_charts_tab(writer, filepath)

        # 8. Export companion interactive HTML files
        if export_html:
            self._export_html_charts(filepath)

        logger.info(f"Analysis results exported to: {filepath}")

    def _write_summary_tab(self, writer: pd.ExcelWriter, format_cells: bool) -> None:
        """Write summary tab with analysis metadata."""
        # Convert summary dict to DataFrame for better Excel formatting
        summary_items = []

        # SDS Information
        summary_items.append(('Category', 'Sampling Design State'))
        summary_items.append(('SDS', self.result.summary['sds']))
        summary_items.append(('Description', self.result.summary['sds_description']))
        summary_items.append(('Replication Type', self.result.summary['replication_type']))
        summary_items.append(('', ''))

        # Analysis Configuration
        summary_items.append(('Category', 'Analysis Configuration'))
        summary_items.append(('Analysis Type', self.result.summary['analysis_type']))
        summary_items.append(('Response Variable', self.result.summary['response_var']))
        summary_items.append(('Grouping Variables', str(self.result.summary['grouping_vars'])))
        summary_items.append(('Time Variable', self.result.summary['time_var']))
        summary_items.append(('', ''))

        # Data Dimensions
        summary_items.append(('Category', 'Data Dimensions'))
        summary_items.append(('Total Observations', self.result.summary['n_observations']))
        summary_items.append(('Number of Charts', self.result.summary['n_charts']))
        summary_items.append(('Chart Types', ', '.join(self.result.summary['chart_types'])))
        summary_items.append(('Is Stratified', self.result.summary['is_stratified']))
        summary_items.append(('', ''))

        # Capabilities
        summary_items.append(('Category', 'Analysis Capabilities'))
        summary_items.append(('Has Residuals', self.result.summary['has_residuals']))
        summary_items.append(('Has Effects', self.result.summary['has_effects']))
        summary_items.append(('Has Interactions', self.result.summary['has_interactions']))
        summary_items.append(('Variance Decomposition', self.result.summary['variance_decomposition']))
        summary_items.append(('Interaction Analysis', self.result.summary['interaction_analysis']))
        summary_items.append(('', ''))

        # Signals
        summary_items.append(('Category', 'Process Signals'))
        summary_items.append(('Total Signals Detected', self.result.summary['n_signals_total']))

        # Add SDS capabilities as a list
        if self.result.summary['sds_capabilities']:
            summary_items.append(('', ''))
            summary_items.append(('Category', 'SDS Capabilities'))
            for capability in self.result.summary['sds_capabilities']:
                summary_items.append(('', capability))

        # Create DataFrame
        summary_df = pd.DataFrame(summary_items, columns=['Attribute', 'Value'])

        # Write to Excel
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Apply formatting if requested
        if format_cells:
            ws = writer.sheets['Summary']
            self._apply_formatting(ws)

    def _write_chart_tabs(self, writer: pd.ExcelWriter, format_cells: bool) -> None:
        """Write tabs for each chart."""

        # Check if this is a stratified analysis (XmR/R with grouping_vars)
        if self.result.summary['is_stratified']:
            # Stratified analysis: combine all stratified charts into single tab
            self._write_stratified_chart_tab(writer, format_cells)

            # For stratified XmR/R charts, also create a summary tab for quick comparison
            chart_type = self.result._analysis_type
            if chart_type in ['XmR', 'R']:
                self._write_stratified_summary_tab(writer, format_cells)
        else:
            # Standard analysis: each chart gets its own tab
            for chart_name, chart_info in self.result.charts.items():
                # Get chart data
                chart_data = chart_info.get('data')
                if chart_data is None or not isinstance(chart_data, pd.DataFrame):
                    continue

                # Create tab name (Excel limit: 31 chars)
                tab_name = f"Chart_{chart_name}"
                if len(tab_name) > 31:
                    tab_name = tab_name[:31]

                # Write chart data
                chart_data.to_excel(writer, sheet_name=tab_name, index=False)

                # Apply formatting if requested
                if format_cells:
                    ws = writer.sheets[tab_name]
                    self._apply_formatting(ws)

    def _write_stratified_chart_tab(
        self,
        writer: pd.ExcelWriter,
        format_cells: bool
    ) -> None:
        """
        Write all stratified charts to a single combined tab.

        Combines all strata into one worksheet with a column identifying the stratum,
        making it easy to compare and filter in Excel.
        """
        # Collect all stratified charts
        combined_data = []

        for chart_name, chart_info in self.result.charts.items():
            # Skip standard charts (will be written separately)
            if chart_name in STANDARD_CHART_NAMES:
                continue

            chart_data = chart_info.get('data')
            if chart_data is None or not isinstance(chart_data, pd.DataFrame):
                continue

            # For XmR/R with grouping_vars, the 'rsg' column already exists
            # and contains the stratification identifier (chart name)
            # No need to add it - just use the data as-is
            combined_data.append(chart_data.copy())

        # Combine all stratified charts
        if combined_data:
            combined_df = pd.concat(combined_data, ignore_index=True)

            # Create descriptive tab name
            chart_type = self.result._analysis_type
            grouping_vars = self.result._ads.spec.rsg_vars if self.result._ads.spec.has_grouping else []

            if len(grouping_vars) == 1:
                tab_name = f"Chart_{chart_type}_by_{grouping_vars[0]}"
            else:
                tab_name = f"Chart_{chart_type}_Stratified"

            # Excel tab name limit: 31 chars
            if len(tab_name) > 31:
                tab_name = tab_name[:31]

            # Write combined data
            combined_df.to_excel(writer, sheet_name=tab_name, index=False)

            # Apply formatting if requested
            if format_cells:
                ws = writer.sheets[tab_name]
                self._apply_formatting(ws)

        # Also write any standard charts (Xbar, S, etc.)
        # These are separate analyses, not stratified
        for chart_name in STANDARD_CHART_NAMES:
            if chart_name in self.result.charts:
                chart_info = self.result.charts[chart_name]
                chart_data = chart_info.get('data')
                if chart_data is not None and isinstance(chart_data, pd.DataFrame):
                    tab_name = f"Chart_{chart_name}"
                    chart_data.to_excel(writer, sheet_name=tab_name, index=False)

                    if format_cells:
                        ws = writer.sheets[tab_name]
                        self._apply_formatting(ws)

    def _write_stratified_summary_tab(
        self,
        writer: pd.ExcelWriter,
        format_cells: bool
    ) -> None:
        """
        Create a summary tab for stratified XmR/R charts.

        Provides a high-level comparison across all strata showing:
        - Stratum identifier (RSG)
        - Number of observations per stratum
        - Mean, LPL, UPL for each stratum
        - Total signals (beyond_limits != 0)
        - Signal rate (% of observations with signals)
        """
        # Collect data from all stratified charts
        all_chart_data = []

        for chart_name, chart_info in self.result.charts.items():
            # Skip standard charts
            if chart_name in STANDARD_CHART_NAMES:
                continue

            chart_data = chart_info.get('data')
            if chart_data is not None and isinstance(chart_data, pd.DataFrame):
                all_chart_data.append(chart_data)

        # Combine all chart data
        if not all_chart_data:
            return

        combined_data = pd.concat(all_chart_data, ignore_index=True)

        # Group by stratum (rsg column) and calculate summary statistics
        if 'rsg' not in combined_data.columns:
            return

        summary_rows = []

        for stratum, group in combined_data.groupby('rsg', sort=False, observed=True):
            n_obs = len(group)

            # Get mean, lpl, upl (should be constant within stratum)
            mean_val = group['mean'].iloc[0] if 'mean' in group.columns else None
            lpl_val = group['lpl'].iloc[0] if 'lpl' in group.columns else None
            upl_val = group['upl'].iloc[0] if 'upl' in group.columns else None

            # Count signals
            if 'beyond_limits' in group.columns:
                n_signals = (group['beyond_limits'] != 0).sum()
                signal_rate = (n_signals / n_obs * 100) if n_obs > 0 else 0
            else:
                n_signals = None
                signal_rate = None

            summary_rows.append({
                'Stratum': stratum,
                'Observations': n_obs,
                'Mean': mean_val,
                'LPL': lpl_val,
                'UPL': upl_val,
                'Signals': n_signals,
                'Signal_Rate_%': signal_rate
            })

        if summary_rows:
            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_rows)

            # Sort by signal count (descending) so worst strata appear first
            if 'Signals' in summary_df.columns:
                summary_df = summary_df.sort_values('Signals', ascending=False)

            # Write to Excel
            tab_name = 'Stratified_Summary'
            summary_df.to_excel(writer, sheet_name=tab_name, index=False)

            if format_cells:
                ws = writer.sheets[tab_name]
                self._apply_formatting(ws)

    def _write_residuals_tab(self, writer: pd.ExcelWriter, format_cells: bool) -> None:
        """Write residuals tab with original data columns."""
        if self.result.residuals is not None:
            # Include original data columns with residuals for context
            # Get original input columns from summary
            time_var = self.result.summary.get('time_var')
            grouping_vars = self.result.summary.get('grouping_vars', [])
            response_var = self.result.summary.get('response_var')

            # Build list of columns to include: original inputs + residuals
            original_cols = []
            if time_var and time_var in self.result.dataset.columns:
                original_cols.append(time_var)
            if grouping_vars:
                original_cols.extend([g for g in grouping_vars if g in self.result.dataset.columns])
            if response_var and response_var in self.result.dataset.columns:
                original_cols.append(response_var)

            # Add residual columns
            residual_cols = ['R1', 'R2', 'R3', 'R4', 'R5']
            all_cols = original_cols + residual_cols

            # Extract these columns from dataset (which has both)
            residuals_with_data = self.result.dataset[all_cols].copy()

            # Write to Excel
            residuals_with_data.to_excel(writer, sheet_name='Residuals', index=False)

            if format_cells:
                ws = writer.sheets['Residuals']
                self._apply_formatting(ws)

    def _write_effects_tab(self, writer: pd.ExcelWriter, format_cells: bool) -> None:
        """Write effects tab."""
        # Convert effects dict to DataFrame
        effects_data = []

        for effect_name, effect_values in self.result.effects.items():
            if isinstance(effect_values, pd.DataFrame):
                # For DataFrames, extract all rows
                for _idx, row in effect_values.iterrows():
                    # Get the value column - look for specific patterns
                    # Exclude first column (grouping variable) and match effect/ME columns
                    value_col = [
                        c for c in effect_values.columns[1:]
                        if 'effect' in c.lower()
                        or c.upper().endswith('_ME')
                        or c.upper() == 'PT_ME'
                    ]
                    # Use first match or last column as value
                    val = row[value_col[0]] if value_col else row.iloc[-1]

                    # Get the level/category
                    level_col = effect_values.columns[0]
                    level = row[level_col]

                    effects_data.append({
                        'Effect_Type': effect_name,
                        'Level': str(level),
                        'Value': val
                    })
            elif isinstance(effect_values, pd.Series):
                for idx, val in effect_values.items():
                    effects_data.append({
                        'Effect_Type': effect_name,
                        'Level': str(idx),
                        'Value': val
                    })
            elif isinstance(effect_values, (int, float)):
                effects_data.append({
                    'Effect_Type': effect_name,
                    'Level': '',
                    'Value': effect_values
                })

        if effects_data:
            effects_df = pd.DataFrame(effects_data)
            effects_df.to_excel(writer, sheet_name='Effects', index=False)

            if format_cells:
                ws = writer.sheets['Effects']
                self._apply_formatting(ws)

    def _write_interactions_tab(self, writer: pd.ExcelWriter, format_cells: bool) -> None:
        """Write interactions tab."""
        # Convert interactions dict to DataFrame
        interactions_data = []

        for interaction_name, interaction_values in self.result.interactions.items():
            if isinstance(interaction_values, pd.Series):
                # Handle MultiIndex Series (cell-level interactions)
                if isinstance(interaction_values.index, pd.MultiIndex):
                    # Extract index level names and values
                    for idx_tuple, val in interaction_values.items():
                        # Build a dict with index level values plus the interaction value
                        row_dict = {'Interaction': interaction_name}

                        # Add each index level as a separate column
                        for level_name, level_val in zip(interaction_values.index.names, idx_tuple):
                            row_dict[level_name] = level_val

                        row_dict['Value'] = val
                        interactions_data.append(row_dict)
                else:
                    # Regular index - use as Combination
                    for idx, val in interaction_values.items():
                        interactions_data.append({
                            'Interaction': interaction_name,
                            'Combination': str(idx),
                            'Value': val
                        })
            elif isinstance(interaction_values, pd.DataFrame):
                # For DataFrames, flatten them
                for row_idx, row in interaction_values.iterrows():
                    for col_name, val in row.items():
                        interactions_data.append({
                            'Interaction': interaction_name,
                            'Combination': f"{row_idx} x {col_name}",
                            'Value': val
                        })

        if interactions_data:
            interactions_df = pd.DataFrame(interactions_data)
            interactions_df.to_excel(writer, sheet_name='Interactions', index=False)

            if format_cells:
                ws = writer.sheets['Interactions']
                self._apply_formatting(ws)

    def _write_full_dataset_tab(self, writer: pd.ExcelWriter, format_cells: bool) -> None:
        """Write full dataset tab."""
        self.result.dataset.to_excel(writer, sheet_name='Full_Dataset', index=False)

        if format_cells:
            ws = writer.sheets['Full_Dataset']
            self._apply_formatting(ws)

    def _write_visual_charts_tab(self, writer: pd.ExcelWriter, filepath: str) -> None:
        """
        Write visual charts tab with embedded chart images.

        Creates a 'Visual_Charts' worksheet with:
        - Embedded PNG images of control charts
        - Hyperlinks to interactive HTML files
        - Organized layout for presentation
        """
        try:
            from io import BytesIO

            from openpyxl.drawing.image import Image
            from openpyxl.styles import Font

            from ..plotting import Plotter

            # Get workbook to add images
            wb = writer.book
            ws = wb.create_sheet('Visual_Charts')

            # Create plotter
            plotter = Plotter(self.result)

            # Track row position for layout
            current_row = 1

            # Export combined charts first (Xbar, S, etc.)
            combined_charts = [name for name in self.result.charts
                             if name in STANDARD_CHART_NAMES]

            if combined_charts:
                ws[f'A{current_row}'] = 'COMBINED CONTROL CHARTS'
                ws[f'A{current_row}'].font = Font(bold=True, size=14)
                current_row += 2

                for chart_name in combined_charts:
                    try:
                        # Generate chart
                        fig = plotter.plot(
                            chart=chart_name,
                            width=1200,
                            height=500,
                            template='processbehavior'
                        )

                        # Save as image
                        img_buffer = BytesIO()
                        fig.figure.write_image(img_buffer, format='png', width=1200, height=500)
                        img_buffer.seek(0)

                        # Add title
                        ws[f'A{current_row}'] = f"{chart_name} Chart"
                        ws[f'A{current_row}'].font = Font(bold=True)
                        current_row += 1

                        # Add image
                        img = Image(img_buffer)
                        img.width = 800  # Scale to fit Excel
                        img.height = 333
                        ws.add_image(img, f'A{current_row}')
                        current_row += 18  # Space for image + margin

                        logger.info(f"Added {chart_name} chart image to Excel")

                    except Exception as e:
                        logger.warning(f"Could not add {chart_name} chart image: {e}")
                        ws[f'A{current_row}'] = f"Error generating {chart_name} chart"
                        current_row += 2

            # Add note about interactive HTML files
            current_row += 2
            ws[f'A{current_row}'] = 'INTERACTIVE CHARTS'
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            current_row += 1
            ws[f'A{current_row}'] = 'Interactive HTML files have been exported alongside this Excel file.'
            current_row += 1
            ws[f'A{current_row}'] = (
                'Open the .html files in a web browser for full interactivity '
                '(zoom, pan, hover tooltips).'
            )

        except ImportError as e:
            logger.warning(f"Could not create visual charts tab: {e}")
            logger.warning("Install kaleido for image export: pip install kaleido")

    def _export_html_charts(self, filepath: str) -> None:
        """
        Export interactive HTML charts alongside the Excel file.

        Creates HTML files in the same directory as the Excel file:
        - {basename}_combined.html - Combined Xbar/S charts
        - {basename}_stratified.html - Stratified XmR charts (if applicable)
        """
        try:
            from pathlib import Path

            from ..plotting import Plotter

            output_path = Path(filepath)
            output_dir = output_path.parent
            base_name = output_path.stem

            plotter = Plotter(self.result)

            # Export combined charts
            combined_charts = [name for name in self.result.charts
                             if name in STANDARD_CHART_NAMES]

            if combined_charts:
                html_file = output_dir / f'{base_name}_combined.html'
                fig = plotter.plot(
                    width=1400,
                    height=800,
                    template='processbehavior',
                    title=f'Control Charts: {base_name}'
                )
                fig.save_html(str(html_file))
                logger.info(f"Exported interactive combined charts to: {html_file}")

            # Export stratified charts if present
            stratified_charts = [name for name in self.result.charts
                               if name not in combined_charts and len(self.result.charts) > len(combined_charts)]

            if stratified_charts and self.result.summary.get('is_stratified', False):
                html_file = output_dir / f'{base_name}_stratified.html'
                # This will plot all stratified charts
                fig = plotter.plot(
                    width=1800,
                    height=1200,
                    template='processbehavior',
                    title=f'Stratified Control Charts: {base_name}'
                )
                fig.save_html(str(html_file))
                logger.info(f"Exported interactive stratified charts to: {html_file}")

        except Exception as e:
            logger.warning(f"Could not export HTML charts: {e}")

    def _apply_formatting(self, worksheet) -> None:
        """
        Apply standard formatting to worksheet.

        - Bold headers
        - Freeze top row
        - Auto-size columns
        """
        from openpyxl.styles import Font

        # Bold headers (first row)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Freeze top row
        worksheet.freeze_panes = 'A2'

        # Auto-size columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:  # noqa: S110
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
